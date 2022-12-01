#!/usr/bin/env python
# coding: utf-8

# In[1]:

import pandas as pd
from openpyxl import Workbook, load_workbook
wb = load_workbook("CarData.xlsx")
ws = wb.active
file = pd.read_excel("CarData.xlsx")
file


# In[24]:

dict2 = {1:"A", 2:"B", 3: "C", 4:"D", 5 : "E", 6 : "F", 7 : "G"}
#Created a dictionary to store all the states
dict_states={}
dict_statenum={}
for i in range(23):       
    dict_states[file.iloc[i,0]]= i
    dict_statenum[i]=file.iloc[i,0]

#find the highest value in a colum and tells you what row that highest number belongs to
def Hi(x):
    b=0
    CurrentRow=0
    Row=0
    Highest=0
    #iterates through all values in colum 
    for i in ws.iter_rows(min_row = 2,
                         max_row = 24,
                         min_col=x,
                         max_col=x,
                         values_only=True):
        
        #makes string of the value in cell by default iter_rows outputs cell calue fo 12.6 as (12.6,)
        L=str(i)
        #num is a string that will be used to store the cell value with out the added () and , 
        num=""
        #strips cell value of everything other then the number
        for n in L:
            try: 
                #crow tracks which row the cell we are looking at is in
                crow=int(n)
                num=num+n
                #print(thing)

            except ValueError:
                if n == ".":
                    num=num+n
                    #print(thing)
                pass
        if num == "":
            num="0"
        #coverts num into float
        num=float(num)
        #checks if new number is higher
        if Highest < num:
            Highest=num
            Row=CurrentRow
        CurrentRow+=1
    state=dict_statenum[Row]
    category=ws[str(dict2[x])+"1"].value
    print("The highest number in the category: ",category," is:",Highest," which belongs to: ",state)

for i in range(6):
    Hi(i+2)
    #Lo(i+2)    
class state():
    #User gets to pick a state 
    def __init__(self, name):
        self.name = name
        self.num = 0
    
    def getAllValue(self):
        #Print all value in the row of the state picked
        print(file.iloc[dict_states[self.name]])
    

    def getValue(self, index):
        indent = "\n" + " "*23
        #User can choose to retrive only one value. Must enter an integer to corresponding column
        try:
            print(file.iloc[dict_states[self.name], [index]])
        except IndexError:
            print("Error: Enter a number. 1 : Age 0-20" + indent + "2 : Age 21-3" + indent + "3 : Age 35-54" + indent + \
                 "4 : 55+" + indent + "5 : All ages" + indent + "6 : Male" + indent + "7 : Female")
            
    def setValue(self, col, newValue):
        #Created new dictionary to retrieve columns when user enter integer
        dict2 = {1:"A", 2:"B", 3: "C", 4:"D", 5 : "E", 6 : "F", 7 : "G"}
        #When enter integer, this code automatically assigns it to a cell in excel. For example, if enter "1" and state is Delaware, the cell is D9
        #We are adding "2" because the rows start 1 on the headers
        ws[str(dict2[col])+str(dict_states[self.name] + 2)] = newValue
        wb.save("CarData.xlsx")
        print(ws["D9"].value)
        
minn = state("Delaware")
minn.getAllValue()
minn.getValue("8")
minn.setValue(4, 3.0)


                


                


                






